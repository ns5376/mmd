#!/usr/bin/env python3
"""
Add WAAMD Author ID matches to a pipeline CSV.

Matching strategy (in order):
  1. Exact normalized Arabic match  (author col vs Ar Name)
  2. Token-sort Arabic match        (order-insensitive)
  3. Fuzzy Arabic similarity        (SequenceMatcher on normalized Arabic)
  4. Fallback: fuzzy English match  (romanized name vs Eng Name)

Reads a pipeline-style CSV, matches author names against `waamd.csv`,
and writes a new CSV with a `WAAMD Author ID` column added.
"""

import argparse
import csv
import os
import re
import unicodedata
from difflib import SequenceMatcher
from collections import defaultdict


BASE_DIR = os.path.dirname(os.path.abspath(__file__))
_csv_path = os.path.join(BASE_DIR, "waamd.csv")
_xlsx_path = os.path.join(BASE_DIR, "waamd.xlsx")
DEFAULT_WAAMD_PATH = _csv_path if os.path.exists(_csv_path) else _xlsx_path
UTF8 = "utf-8"
UTF8_SIG = "utf-8-sig"


# ---------------------------------------------------------------------------
# Arabic normalization
# ---------------------------------------------------------------------------

def normalize_arabic(text: str) -> str:
    """Normalize Arabic text for comparison: remove diacritics, normalize alef/ya/ta-marbuta."""
    if not text:
        return ""
    text = str(text).strip()
    # Remove tatweel
    text = text.replace("\u0640", "")
    # Remove diacritics (harakat + shadda + tanwin)
    diacritics = "\u064b\u064c\u064d\u064e\u064f\u0650\u0651\u0652\u0653\u0654\u0655\u0656\u0657\u0658"
    for ch in diacritics:
        text = text.replace(ch, "")
    # Normalize alef variants → bare alef
    text = re.sub(r"[\u0622\u0623\u0625\u0671]", "\u0627", text)
    # Normalize ya: alef maqsura → ya
    text = text.replace("\u0649", "\u064a")
    # Normalize ta marbuta → ha
    text = text.replace("\u0629", "\u0647")
    # Collapse whitespace
    text = re.sub(r"\s+", " ", text).strip()
    return text


def arabic_token_sort(text: str) -> str:
    tokens = normalize_arabic(text).split()
    tokens.sort()
    return " ".join(tokens)


# ---------------------------------------------------------------------------
# English normalization (for fallback)
# ---------------------------------------------------------------------------

def strip_accents(text: str) -> str:
    text = unicodedata.normalize("NFKD", text)
    return "".join(ch for ch in text if not unicodedata.combining(ch))


def normalize_english(text: str) -> str:
    if not text:
        return ""
    text = str(text).strip().lower()
    text = strip_accents(text)
    text = text.replace("\u2018", "'").replace("\u2019", "'").replace("`", "'")
    text = text.replace("\u02bf", "'").replace("\u02be", "'")
    text = text.replace("b.", " ibn ")
    text = text.replace(" bin ", " ibn ").replace(" ben ", " ibn ")
    text = re.sub(r"[\.,\/\-\_\(\)\[\]]+", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def token_sort_key(text: str) -> str:
    tokens = [t for t in normalize_english(text).split() if t]
    tokens.sort()
    return " ".join(tokens)


def similarity(a: str, b: str) -> float:
    if not a or not b:
        return 0.0
    return SequenceMatcher(None, a, b).ratio()


# ---------------------------------------------------------------------------
# WAAMD loading
# ---------------------------------------------------------------------------

def load_waamd_authors(path: str):
    if path.lower().endswith(".csv"):
        return _load_waamd_from_csv(path)
    return _load_waamd_from_xlsx(path)


def _load_waamd_from_csv(csv_path: str):
    records = []
    with open(csv_path, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            waamd_id = str(row.get("\u067e") or row.get("id") or "").strip()
            eng_name = str(row.get("Eng Name") or "").strip()
            ar_name = str(row.get("Ar Name") or "").strip()
            if not waamd_id or not eng_name:
                continue
            if waamd_id.endswith(".0") and waamd_id[:-2].isdigit():
                waamd_id = waamd_id[:-2]
            records.append(_make_record(waamd_id, eng_name, ar_name))
    return records


def _load_waamd_from_xlsx(excel_path: str):
    from openpyxl import load_workbook
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    if "WAAMD Authors" not in wb.sheetnames:
        raise RuntimeError(f"'WAAMD Authors' sheet not found in {excel_path}")
    ws = wb["WAAMD Authors"]
    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        waamd_id = "" if row[0] is None else str(row[0]).strip()
        eng_name = "" if len(row) < 2 or row[1] is None else str(row[1]).strip()
        ar_name = "" if len(row) < 3 or row[2] is None else str(row[2]).strip()
        if not waamd_id or not eng_name:
            continue
        if waamd_id.endswith(".0") and waamd_id[:-2].isdigit():
            waamd_id = waamd_id[:-2]
        records.append(_make_record(waamd_id, eng_name, ar_name))
    return records


def _make_record(waamd_id, eng_name, ar_name):
    return {
        "waamd_id": waamd_id,
        "eng_name": eng_name,
        "ar_name": ar_name,
        "ar_norm": normalize_arabic(ar_name),
        "ar_token_sort": arabic_token_sort(ar_name),
        "eng_norm": normalize_english(eng_name),
        "eng_token_sort": token_sort_key(eng_name),
    }


# ---------------------------------------------------------------------------
# Index building
# ---------------------------------------------------------------------------

def build_waamd_indexes(waamd_records):
    by_ar_norm = {}
    by_ar_sort = {}
    by_eng_norm = {}
    by_eng_sort = {}
    by_ar_token = defaultdict(list)
    by_eng_token = defaultdict(list)

    for rec in waamd_records:
        by_ar_norm.setdefault(rec["ar_norm"], rec)
        by_ar_sort.setdefault(rec["ar_token_sort"], rec)
        by_eng_norm.setdefault(rec["eng_norm"], rec)
        by_eng_sort.setdefault(rec["eng_token_sort"], rec)
        for tok in rec["ar_norm"].split():
            by_ar_token[tok].append(rec)
        for tok in rec["eng_norm"].split():
            by_eng_token[tok].append(rec)

    return {
        "by_ar_norm": by_ar_norm,
        "by_ar_sort": by_ar_sort,
        "by_eng_norm": by_eng_norm,
        "by_eng_sort": by_eng_sort,
        "by_ar_token": by_ar_token,
        "by_eng_token": by_eng_token,
    }


# ---------------------------------------------------------------------------
# Matching
# ---------------------------------------------------------------------------

def _candidate_set(tokens, token_index, waamd_records, max_candidates=80):
    """Return candidate records that share at least one token with query."""
    overlap = defaultdict(int)
    for tok in set(tokens):
        for rec in token_index.get(tok, []):
            overlap[rec["waamd_id"]] += 1
    if not overlap:
        return waamd_records
    top_ids = {wid for wid, _ in sorted(overlap.items(), key=lambda x: x[1], reverse=True)[:max_candidates]}
    seen = set()
    out = []
    for rec in waamd_records:
        if rec["waamd_id"] in top_ids and rec["waamd_id"] not in seen:
            seen.add(rec["waamd_id"])
            out.append(rec)
    return out


def best_waamd_match(ar_name: str, eng_name: str, waamd_records, idx, threshold: float):
    """
    Try to match using Arabic first, English as fallback.
    Returns (waamd_id, matched) — no score/method exposed.
    """
    ar_norm = normalize_arabic(ar_name)
    ar_sort = arabic_token_sort(ar_name)
    eng_norm = normalize_english(eng_name)
    eng_sort = token_sort_key(eng_name)

    # --- Arabic exact ---
    if ar_norm:
        if ar_norm in idx["by_ar_norm"]:
            return idx["by_ar_norm"][ar_norm]["waamd_id"]
        if ar_sort in idx["by_ar_sort"]:
            return idx["by_ar_sort"][ar_sort]["waamd_id"]

    # --- English exact ---
    if eng_norm:
        if eng_norm in idx["by_eng_norm"]:
            return idx["by_eng_norm"][eng_norm]["waamd_id"]
        if eng_sort in idx["by_eng_sort"]:
            return idx["by_eng_sort"][eng_sort]["waamd_id"]

    # --- Arabic fuzzy ---
    best_id = ""
    best_score = 0.0
    if ar_norm:
        ar_tokens = ar_norm.split()
        candidates = _candidate_set(ar_tokens, idx["by_ar_token"], waamd_records)
        for rec in candidates:
            if not rec["ar_norm"]:
                continue
            if abs(len(ar_norm) - len(rec["ar_norm"])) > 20:
                continue
            score = max(
                similarity(ar_norm, rec["ar_norm"]),
                similarity(ar_sort, rec["ar_token_sort"]),
            )
            if score > best_score:
                best_score = score
                best_id = rec["waamd_id"]
        if best_score >= threshold:
            return best_id

    # --- English fuzzy fallback ---
    best_id = ""
    best_score = 0.0
    if eng_norm:
        eng_tokens = eng_norm.split()
        candidates = _candidate_set(eng_tokens, idx["by_eng_token"], waamd_records)
        for rec in candidates:
            if abs(len(eng_norm) - len(rec["eng_norm"])) > 35:
                continue
            score = max(
                similarity(eng_norm, rec["eng_norm"]),
                similarity(eng_sort, rec["eng_token_sort"]),
            )
            if score > best_score:
                best_score = score
                best_id = rec["waamd_id"]
        if best_score >= threshold:
            return best_id

    return ""


# ---------------------------------------------------------------------------
# CSV I/O
# ---------------------------------------------------------------------------

def choose_columns(fieldnames):
    """Return (arabic_author_col, english_author_col). Either may be None."""
    ar_col = "author" if "author" in fieldnames else None
    eng_candidates = ["AUTHOR NAME ENGLISH", "Author Name - English", "Author Name - Eng.", "author_name_english"]
    eng_col = next((c for c in eng_candidates if c in fieldnames), None)
    if not ar_col and not eng_col:
        raise RuntimeError(
            "Could not find an author column. Expected 'author' (Arabic) and/or one of: " + ", ".join(eng_candidates)
        )
    return ar_col, eng_col


def read_csv_rows(path: str):
    with open(path, "r", encoding=UTF8_SIG, newline="") as f:
        reader = csv.DictReader(f)
        rows = list(reader)
        return reader.fieldnames or [], rows


def write_csv_rows(path: str, fieldnames, rows):
    with open(path, "w", encoding=UTF8_SIG, newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def add_waamd_ids(input_csv: str, output_csv: str, waamd_path: str, threshold: float):
    fieldnames, rows = read_csv_rows(input_csv)
    if not fieldnames:
        raise RuntimeError(f"No CSV headers found in {input_csv}")

    ar_col, eng_col = choose_columns(fieldnames)
    waamd_records = load_waamd_authors(waamd_path)
    if not waamd_records:
        raise RuntimeError(f"No WAAMD author rows found in {waamd_path}")
    idx = build_waamd_indexes(waamd_records)

    cache = {}
    match_count = 0
    for row in rows:
        ar_name = row.get(ar_col, "") if ar_col else ""
        eng_name = row.get(eng_col, "") if eng_col else ""
        cache_key = (ar_name, eng_name)
        if cache_key not in cache:
            cache[cache_key] = best_waamd_match(ar_name, eng_name, waamd_records, idx, threshold)
        waamd_id = cache[cache_key]
        row["WAAMD Author ID"] = waamd_id
        if waamd_id:
            match_count += 1

    # Insert WAAMD Author ID right after the Arabic author column (or English if no Arabic)
    output_fields = list(fieldnames)
    if "WAAMD Author ID" not in output_fields:
        anchor = ar_col or eng_col
        try:
            insert_at = output_fields.index(anchor) + 1
        except ValueError:
            insert_at = len(output_fields)
        output_fields.insert(insert_at, "WAAMD Author ID")

    write_csv_rows(output_csv, output_fields, rows)
    return len(rows), match_count, ar_col or eng_col


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def default_output_path(input_csv: str) -> str:
    root, ext = os.path.splitext(input_csv)
    return f"{root}_with_waamd_ids{ext or '.csv'}"


def main():
    parser = argparse.ArgumentParser(
        description="Match pipeline CSV authors to waamd.csv and add WAAMD Author ID."
    )
    parser.add_argument("input_csv", help="Input pipeline CSV")
    parser.add_argument("--output", help="Output CSV path")
    parser.add_argument("--waamd", default=DEFAULT_WAAMD_PATH)
    parser.add_argument("--threshold", type=float, default=0.82,
                        help="Minimum similarity score 0.0–1.0 (default: 0.82)")
    args = parser.parse_args()

    output_csv = args.output or default_output_path(args.input_csv)
    total, matched, col = add_waamd_ids(
        input_csv=args.input_csv,
        output_csv=output_csv,
        waamd_path=args.waamd,
        threshold=args.threshold,
    )
    print(f"Author column: {col}")
    print(f"Rows processed: {total}")
    print(f"WAAMD IDs matched: {matched}")
    print(f"Saved: {output_csv}")


if __name__ == "__main__":
    main()
